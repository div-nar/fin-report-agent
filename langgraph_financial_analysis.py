import os
import pandas as pd
from typing import TypedDict, List, Dict
from langgraph.graph import StateGraph, END
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_core.messages import HumanMessage, SystemMessage
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import time

# State definition
class FinancialAnalysisState(TypedDict):
    raw_transactions: List[Dict]
    categorized_transactions: List[Dict]
    capex_total: float
    opex_total: float
    capex_by_category: Dict[str, float]
    opex_by_category: Dict[str, float]
    report_path: str
    errors: List[str]
    stats: Dict[str, int]

# Initialize LLM - Gemini 2.5 Flash
def get_llm():
    api_key = os.getenv('GOOGLE_API_KEY') or os.getenv('GEMINI_API_KEY')
    if not api_key:
        raise ValueError("Please set GOOGLE_API_KEY environment variable")
    return ChatGoogleGenerativeAI(
        model="gemini-2.5-flash",
        google_api_key=api_key,
        temperature=0.1,
        max_retries=2,
        request_timeout=60
    )

# Node 1: Load data with FILTERING
def load_data(state: FinancialAnalysisState) -> FinancialAnalysisState:
    print("\n[1/4] Loading and filtering data...")
    
    transactions = []
    
    # Kodo-Pay Sheet
    try:
        kodo_df = pd.read_csv("kodo-pay-reimbursement.csv")
        kodo_debit = kodo_df[kodo_df['Dr/Cr'] == 'Dr'].copy()
        
        for _, row in kodo_debit.iterrows():
            date = str(row['Date (IST)']) if pd.notna(row['Date (IST)']) else ''
            amount = float(row['Txn Amount (INR)']) if pd.notna(row['Txn Amount (INR)']) else 0.0
            narration = str(row['Narration on Kodo Pay']) if pd.notna(row['Narration on Kodo Pay']) else ''
            category = str(row['Category']) if pd.notna(row['Category']) else 'Uncategorized'
            comments = str(row['Maker Comments']) if pd.notna(row['Maker Comments']) else ''
            maker_name = str(row['Maker Name']) if pd.notna(row['Maker Name']) else ''
            status = str(row['Outward Payment Status']) if pd.notna(row['Outward Payment Status']) else ''
            
            # FILTER: Skip LQ Prepaid transactions (internal transfers)
            if 'lq prepaid' in narration.lower() or 'lq prepaid' in comments.lower():
                continue
            
            full_description = f"{narration} | {category} | {comments} | {maker_name} | {status}"
            
            transactions.append({
                'source': 'Kodo-Pay',
                'date': date,
                'amount': amount,
                'category': category,
                'narration': narration,
                'comments': comments,
                'maker_name': maker_name,
                'description': full_description
            })
        print(f"  ✓ Kodo-Pay: {len(transactions)} transactions (filtered out LQ Prepaid)")
    except Exception as e:
        state['errors'].append(f"Kodo-Pay error: {str(e)}")
    
    # Transactions Sheet
    try:
        trans_df = pd.read_csv("transactions-csv.csv")
        trans_debit = trans_df[~trans_df['Txn Category'].isin(['FUNDING', 'CARD_CREDIT'])].copy()
        
        initial_count = len(transactions)
        
        for _, row in trans_debit.iterrows():
            merchant = str(row['Merchant/Narration']) if pd.notna(row['Merchant/Narration']) else ''
            
            # FILTER: Skip LQ Prepaid transactions (internal transfers)
            if 'lq prepaid' in merchant.lower():
                continue
            
            cardholder = f"{row['Cardholder First Name']} {row['Cardholder Last Name']}" if pd.notna(row['Cardholder First Name']) else ''
            date = str(row['Txn Date']) if pd.notna(row['Txn Date']) else ''
            txn_category = str(row['Txn Category']) if pd.notna(row['Txn Category']) else ''
            amount = float(row['Txn Amount (Rs.)']) if pd.notna(row['Txn Amount (Rs.)']) else 0.0
            category = str(row['Expense Category']) if pd.notna(row['Expense Category']) else 'Uncategorized'
            notes = str(row['Notes']) if pd.notna(row['Notes']) else ''
            
            full_description = f"{merchant} | {category} | {notes} | {cardholder} | {txn_category}"
            
            transactions.append({
                'source': 'Transactions',
                'date': date,
                'amount': amount,
                'category': category,
                'merchant': merchant,
                'narration': merchant,
                'comments': notes,
                'cardholder': cardholder,
                'description': full_description
            })
        
        print(f"  ✓ Transactions: {len(transactions) - initial_count} transactions (filtered out LQ Prepaid)")
    except Exception as e:
        state['errors'].append(f"Transactions error: {str(e)}")
    
    state['raw_transactions'] = transactions
    state['errors'] = state.get('errors', [])
    state['stats'] = {'total': len(transactions), 'llm_categorized': 0, 'uncategorized': 0}
    print(f"  ✓ Total: {len(transactions)} transactions (LQ Prepaid excluded)\n")
    return state

# Node 2: Business-specific rules + LLM categorization
def categorize_with_llm(state: FinancialAnalysisState) -> FinancialAnalysisState:
    print(f"[2/4] Applying business rules + LLM for {len(state['raw_transactions'])} transactions...")
    
    llm = get_llm()
    categorized = []
    llm_needed = []
    
    # BUSINESS-SPECIFIC RULES (applied first)
    print("  → Applying business-specific rules...")
    rule_based_count = 0
    
    for txn in state['raw_transactions']:
        category = txn.get('category', '').lower()
        description = txn.get('description', '').lower()
        narration = txn.get('narration', '').lower()
        comments = txn.get('comments', '').lower()
        maker_name = txn.get('maker_name', '').lower()
        
        all_text = f"{category} {description} {narration} {comments}"
        
        # RULE 1: Mechanical Hardware = CAPEX (hardware company)
        if 'mechanical hardware' in category or 'mechanical hardware' in all_text:
            categorized.append({
                **txn,
                'original_category': txn['category'],
                'expense_type': 'CAPEX',
                'confidence': 'high',
                'reasoning': 'Business Rule: Mechanical Hardware is long-term investment (hardware company)',
                'method': 'business-rule'
            })
            rule_based_count += 1
            continue
        
        # RULE 2: Mister Vishwanatha + Uncategorized + >1000 = CAPEX
        if ('vishwanatha' in maker_name or 'vishwanath' in maker_name) and txn['category'] == 'Uncategorized' and txn['amount'] > 1000:
            categorized.append({
                **txn,
                'original_category': txn['category'],
                'category': 'Capital Investment',  # Assign category
                'expense_type': 'CAPEX',
                'confidence': 'high',
                'reasoning': 'Business Rule: Vishwanatha uncategorized expenses > 1000 → CAPEX (Equipment)',
                'method': 'business-rule'
            })
            rule_based_count += 1
            continue
        
        # Not matched by business rules, needs LLM
        llm_needed.append(txn)
    
    print(f"  ✓ Business rules: {rule_based_count} transactions")
    print(f"  → LLM analysis: {len(llm_needed)} transactions\n")
    
    # LLM categorization for remaining transactions
    if len(llm_needed) == 0:
        state['categorized_transactions'] = categorized
        state['stats']['llm_categorized'] = 0
        state['stats']['business_rules'] = rule_based_count
        return state
    
    # Enhanced prompt with specific business rules
    system_prompt = """You are a financial analyst for a HARDWARE COMPANY.

CRITICAL BUSINESS RULES:
1. ALL "Mechanical Hardware" is CAPEX (since we are a hardware company).
2. If Maker/User is "Vishwanatha" AND Category is "Uncategorized" AND Amount > 1000 -> Assign as CAPEX (Equipment).

Categorization Guidelines:
1. Categorize as CAPEX or OPEX
2. If the category is "Uncategorized", assign a proper category based on the description

CAPEX (Capital Expenditure): One-time investments for long-term use
- Mechanical Hardware (ALWAYS CAPEX), Equipment, machinery, electronics, IT infrastructure, construction
- Examples: Laptops, manufacturing machines, CCTV, office furniture, vehicles

OPEX (Operating Expenditure): Regular/recurring purchases
- Rent, utilities, supplies, food, travel, salaries, maintenance
- Examples: Electricity bills, office supplies, groceries, fuel, hotel stays

Common Categories:
- Food, Grocery, Office Supplies, Utilities, Rent, Travel, Commute
- Electronics, Mechanical Hardware, IT, Lab Work, Construction
- Logistics, Labour, Dog Care, Maintenance

Respond with JSON: {
  "type": "CAPEX" or "OPEX",
  "category": "assigned category name (use existing if not Uncategorized)",
  "reasoning": "brief explanation"
}"""

    batch_size = 5
    delay = 3
    total = len(llm_needed)
    llm_categorized = []
    
    for i in range(0, total, batch_size):
        batch = llm_needed[i:i+batch_size]
        
        batch_prompt = "Categorize and assign proper categories:\n\n"
        for idx, txn in enumerate(batch):
            maker_info = f"Maker: {txn.get('maker_name', 'N/A')}" if txn.get('maker_name') else f"Cardholder: {txn.get('cardholder', 'N/A')}"
            batch_prompt += f"{idx+1}. ₹{txn['amount']:,.2f}\n"
            batch_prompt += f"   Current Category: {txn['category']}\n"
            batch_prompt += f"   {maker_info}\n"
            batch_prompt += f"   Description: {txn['description'][:120]}\n\n"
        
        try:
            messages = [
                SystemMessage(content=system_prompt),
                HumanMessage(content=batch_prompt)
            ]
            
            response = llm.invoke(messages)
            
            content = response.content.strip()
            if content.startswith('```json'):
                content = content.split('```json')[1].split('```')[0].strip()
            elif content.startswith('```'):
                content = content.split('```')[1].split('```')[0].strip()
            
            results = json.loads(content)
            
            for txn, result in zip(batch, results):
                # Use LLM-assigned category if original was Uncategorized
                assigned_category = result.get('category', txn['category'])
                if txn['category'] == 'Uncategorized' and assigned_category != 'Uncategorized':
                    category_note = f" (LLM assigned: {assigned_category})"
                else:
                    category_note = ""
                
                llm_categorized.append({
                    **txn,
                    'category': assigned_category,  # Update category
                    'original_category': txn['category'],  # Keep original
                    'expense_type': result.get('type', 'OPEX'),
                    'confidence': 'llm',
                    'reasoning': f"LLM: {result.get('reasoning', '')}{category_note}",
                    'method': 'pure-llm'
                })
                    
        except Exception as e:
            print(f"  ⚠ Batch failed: {str(e)}, using fallback")
            for txn in batch:
                llm_categorized.append({
                    **txn,
                    'original_category': txn['category'],
                    'expense_type': 'OPEX',
                    'confidence': 'low',
                    'reasoning': f'Error: {str(e)}',
                    'method': 'error-fallback'
                })
        
        current = min(i + batch_size, total)
        print(f"  Progress: {current}/{total}")
        
        if i + batch_size < total:
            time.sleep(delay)
    
    # Combine business-rule and LLM categorized transactions
    all_categorized = categorized + llm_categorized
    state['categorized_transactions'] = all_categorized
    state['stats']['llm_categorized'] = len(llm_categorized)
    state['stats']['business_rules'] = rule_based_count
    
    # Count originally uncategorized that were assigned categories
    originally_uncategorized = len([t for t in all_categorized if t.get('original_category') == 'Uncategorized'])
    still_uncategorized = len([t for t in all_categorized if t['category'] == 'Uncategorized'])
    assigned_categories = originally_uncategorized - still_uncategorized
    
    state['stats']['uncategorized_original'] = originally_uncategorized
    state['stats']['uncategorized_assigned'] = assigned_categories
    state['stats']['uncategorized_remaining'] = still_uncategorized
    
    print(f"  ✓ Complete")
    print(f"  ✓ Business Rules: {rule_based_count}")
    print(f"  ✓ LLM Categorized: {len(llm_categorized)}")
    print(f"  ✓ Originally Uncategorized: {originally_uncategorized}")
    print(f"  ✓ LLM Assigned Categories: {assigned_categories}")
    print(f"  ✓ Still Uncategorized: {still_uncategorized}\n")
    
    return state

# Node 3: Calculate totals
def calculate_totals(state: FinancialAnalysisState) -> FinancialAnalysisState:
    print("[3/4] Calculating...")
    
    capex_total = 0
    opex_total = 0
    capex_by_cat = {}
    opex_by_cat = {}
    
    for txn in state['categorized_transactions']:
        amount = txn['amount']
        category = txn['category']
        
        if txn['expense_type'] == 'CAPEX':
            capex_total += amount
            capex_by_cat[category] = capex_by_cat.get(category, 0) + amount
        else:
            opex_total += amount
            opex_by_cat[category] = opex_by_cat.get(category, 0) + amount
    
    state['capex_total'] = capex_total
    state['opex_total'] = opex_total
    state['capex_by_category'] = capex_by_cat
    state['opex_by_category'] = opex_by_cat
    
    print(f"  ✓ CAPEX: ₹{capex_total:,.2f}")
    print(f"  ✓ OPEX: ₹{opex_total:,.2f}\n")
    
    return state

# Node 4: Generate report
def generate_report(state: FinancialAnalysisState) -> FinancialAnalysisState:
    print("[4/4] Generating report...")
    
    wb = Workbook()
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(bold=True, size=16)
    subtitle_font = Font(bold=True, size=12)
    currency_format = '₹#,##0.00'
    percent_format = '0.0%'
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Executive Summary
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    
    ws_summary['A1'] = "FINANCIAL ANALYSIS REPORT"
    ws_summary['A1'].font = title_font
    ws_summary.merge_cells('A1:D1')
    
    ws_summary['A2'] = f"Period: November 2025"
    ws_summary['A3'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws_summary['A4'] = f"Method: Pure LLM (Gemini 2.5 Flash)"
    
    ws_summary['A6'] = "ANALYSIS NOTES"
    ws_summary['A6'].font = subtitle_font
    ws_summary['A7'] = f"• Total Transactions: {state['stats']['total']}"
    ws_summary['A8'] = f"• Business Rules Applied: {state['stats'].get('business_rules', 0)}"
    ws_summary['A9'] = f"• LLM Categorized: {state['stats']['llm_categorized']}"
    ws_summary['A10'] = f"• Originally Uncategorized: {state['stats'].get('uncategorized_original', 0)}"
    ws_summary['A11'] = f"• LLM Assigned Categories: {state['stats'].get('uncategorized_assigned', 0)}"
    ws_summary['A12'] = f"• Still Uncategorized: {state['stats'].get('uncategorized_remaining', 0)}"
    ws_summary['A13'] = f"• LQ Prepaid: Excluded (internal transfers)"
    
    ws_summary['A15'] = "BUSINESS RULES"
    ws_summary['A15'].font = subtitle_font
    ws_summary['A16'] = "1. Mechanical Hardware → CAPEX (hardware company)"
    ws_summary['A17'] = "2. Mister Vishwanatha + Uncategorized (>1000) → CAPEX"
    
    ws_summary['A19'] = "CATEGORY ASSIGNMENT"
    ws_summary['A19'].font = subtitle_font
    ws_summary['A20'] = "The LLM analyzes transaction descriptions to assign proper categories"
    ws_summary['A21'] = "to items originally marked as 'Uncategorized' in the source data."
    
    ws_summary['A23'] = "EXPENDITURE SUMMARY"
    ws_summary['A23'].font = subtitle_font
    ws_summary.merge_cells('A23:D23')
    
    total_exp = state['capex_total'] + state['opex_total']
    summary_data = [
        ['Type', 'Amount (INR)', '%', 'Count'],
        ['CAPEX', state['capex_total'], state['capex_total']/total_exp if total_exp > 0 else 0,
         len([t for t in state['categorized_transactions'] if t['expense_type'] == 'CAPEX'])],
        ['OPEX', state['opex_total'], state['opex_total']/total_exp if total_exp > 0 else 0,
         len([t for t in state['categorized_transactions'] if t['expense_type'] == 'OPEX'])],
        ['TOTAL', total_exp, 1.0, len(state['categorized_transactions'])]
    ]
    
    for row_idx, row_data in enumerate(summary_data, start=25):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 25:
                cell.fill = header_fill
                cell.font = header_font
            elif row_idx == 28:
                cell.font = Font(bold=True)
            if col_idx == 2 and row_idx > 25:
                cell.number_format = currency_format
            if col_idx == 3 and row_idx > 25:
                cell.number_format = percent_format
            cell.border = border
            cell.alignment = Alignment(horizontal='left' if col_idx == 1 else 'right')
    
    ws_summary.column_dimensions['A'].width = 45
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 12
    
    # CAPEX Statement
    ws_capex = wb.create_sheet("CAPEX Statement")
    ws_capex['A1'] = "CAPITAL EXPENDITURES"
    ws_capex['A1'].font = title_font
    ws_capex.merge_cells('A1:G1')
    
    capex_txns = [t for t in state['categorized_transactions'] if t['expense_type'] == 'CAPEX']
    
    if capex_txns:
        headers = ['Date', 'Source', 'Category', 'Amount', 'Description', 'Method', 'Reasoning']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_capex.cell(row=3, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        for row_idx, txn in enumerate(capex_txns, start=4):
            row_data = [
                txn['date'], txn['source'], txn['category'], txn['amount'],
                txn['description'][:100], txn.get('method', ''), txn['reasoning']
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_capex.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 4:
                    cell.number_format = currency_format
                cell.border = border
                cell.alignment = Alignment(horizontal='left', wrap_text=True)
        
        ws_capex.column_dimensions['A'].width = 12
        ws_capex.column_dimensions['B'].width = 15
        ws_capex.column_dimensions['C'].width = 20
        ws_capex.column_dimensions['D'].width = 18
        ws_capex.column_dimensions['E'].width = 35
        ws_capex.column_dimensions['F'].width = 15
        ws_capex.column_dimensions['G'].width = 40
    
    # OPEX Statement
    ws_opex = wb.create_sheet("OPEX Statement")
    ws_opex['A1'] = "OPERATING EXPENDITURES"
    ws_opex['A1'].font = title_font
    ws_opex.merge_cells('A1:G1')
    
    opex_txns = [t for t in state['categorized_transactions'] if t['expense_type'] == 'OPEX']
    
    if opex_txns:
        headers = ['Date', 'Source', 'Category', 'Amount', 'Description', 'Method', 'Reasoning']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_opex.cell(row=3, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        for row_idx, txn in enumerate(opex_txns, start=4):
            row_data = [
                txn['date'], txn['source'], txn['category'], txn['amount'],
                txn['description'][:100], txn.get('method', ''), txn['reasoning']
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_opex.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 4:
                    cell.number_format = currency_format
                cell.border = border
                cell.alignment = Alignment(horizontal='left', wrap_text=True)
        
        ws_opex.column_dimensions['A'].width = 12
        ws_opex.column_dimensions['B'].width = 15
        ws_opex.column_dimensions['C'].width = 20
        ws_opex.column_dimensions['D'].width = 18
        ws_opex.column_dimensions['E'].width = 35
        ws_opex.column_dimensions['F'].width = 15
        ws_opex.column_dimensions['G'].width = 40
    
    # Category Analysis
    ws_analysis = wb.create_sheet("Category Analysis")
    ws_analysis['A1'] = "CATEGORY BREAKDOWN"
    ws_analysis['A1'].font = title_font
    ws_analysis.merge_cells('A1:C1')
    
    ws_analysis['A3'] = "CAPEX BY CATEGORY"
    ws_analysis['A3'].font = subtitle_font
    ws_analysis.merge_cells('A3:C3')
    
    headers = ['Category', 'Amount', '%']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_analysis.cell(row=4, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    capex_sorted = sorted(state['capex_by_category'].items(), key=lambda x: x[1], reverse=True)
    for idx, (cat, amt) in enumerate(capex_sorted, start=5):
        ws_analysis[f'A{idx}'] = cat
        ws_analysis[f'B{idx}'] = amt
        ws_analysis[f'B{idx}'].number_format = currency_format
        ws_analysis[f'C{idx}'] = amt / state['capex_total'] if state['capex_total'] > 0 else 0
        ws_analysis[f'C{idx}'].number_format = percent_format
        for col in ['A', 'B', 'C']:
            ws_analysis[f'{col}{idx}'].border = border
    
    opex_start = len(capex_sorted) + 7
    ws_analysis[f'A{opex_start}'] = "OPEX BY CATEGORY"
    ws_analysis[f'A{opex_start}'].font = subtitle_font
    ws_analysis.merge_cells(f'A{opex_start}:C{opex_start}')
    
    opex_header_row = opex_start + 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_analysis.cell(row=opex_header_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    opex_sorted = sorted(state['opex_by_category'].items(), key=lambda x: x[1], reverse=True)
    for idx, (cat, amt) in enumerate(opex_sorted, start=opex_header_row+1):
        ws_analysis[f'A{idx}'] = cat
        ws_analysis[f'B{idx}'] = amt
        ws_analysis[f'B{idx}'].number_format = currency_format
        ws_analysis[f'C{idx}'] = amt / state['opex_total'] if state['opex_total'] > 0 else 0
        ws_analysis[f'C{idx}'].number_format = percent_format
        for col in ['A', 'B', 'C']:
            ws_analysis[f'{col}{idx}'].border = border
    
    ws_analysis.column_dimensions['A'].width = 30
    ws_analysis.column_dimensions['B'].width = 20
    ws_analysis.column_dimensions['C'].width = 15
    
    filename = f"Financial_Statement_Hardware_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    state['report_path'] = filename
    
    print(f"  ✓ Saved: {filename}\n")
    return state

# Node 5: Summary
def print_summary(state: FinancialAnalysisState) -> FinancialAnalysisState:
    print("[COMPLETE]")
    print("\n" + "="*80)
    print("HARDWARE COMPANY FINANCIAL ANALYSIS")
    print("="*80)
    print(f"\n📊 Total: {state['stats']['total']} transactions")
    print(f"   Business Rules: {state['stats'].get('business_rules', 0)}")
    print(f"   LLM Categorized: {state['stats']['llm_categorized']}")
    print(f"\n📝 Category Assignment:")
    print(f"   Originally Uncategorized: {state['stats'].get('uncategorized_original', 0)}")
    print(f"   LLM Assigned Categories: {state['stats'].get('uncategorized_assigned', 0)}")
    print(f"   Still Uncategorized: {state['stats'].get('uncategorized_remaining', 0)}")
    print(f"\n💰 CAPEX: ₹{state['capex_total']:,.2f}")
    print(f"   OPEX: ₹{state['opex_total']:,.2f}")
    print(f"   TOTAL: ₹{(state['capex_total'] + state['opex_total']):,.2f}")
    print(f"\n✅ {state['report_path']}")
    print("="*80 + "\n")
    return state

# Build workflow
def build_workflow():
    workflow = StateGraph(FinancialAnalysisState)
    
    workflow.add_node("load_data", load_data)
    workflow.add_node("llm_categorize", categorize_with_llm)
    workflow.add_node("calculate", calculate_totals)
    workflow.add_node("generate_report", generate_report)
    workflow.add_node("summary", print_summary)
    
    workflow.set_entry_point("load_data")
    workflow.add_edge("load_data", "llm_categorize")
    workflow.add_edge("llm_categorize", "calculate")
    workflow.add_edge("calculate", "generate_report")
    workflow.add_edge("generate_report", "summary")
    workflow.add_edge("summary", END)
    
    return workflow.compile()

def main():
    print("\n" + "="*80)
    print("HARDWARE COMPANY FINANCIAL ANALYSIS")
    print("Gemini 2.5 Flash | Hardware Rules | Vishwanatha Logic")
    print("="*80 + "\n")
    
    initial_state = {
        'raw_transactions': [],
        'categorized_transactions': [],
        'capex_total': 0.0,
        'opex_total': 0.0,
        'capex_by_category': {},
        'opex_by_category': {},
        'report_path': '',
        'errors': [],
        'stats': {}
    }
    
    app = build_workflow()
    final_state = app.invoke(initial_state)
    
    if final_state['errors']:
        print("\n⚠️  Errors:")
        for error in final_state['errors']:
            print(f"  - {error}")

if __name__ == "__main__":
    main()
