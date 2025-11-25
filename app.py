import streamlit as st
import pandas as pd
import os
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_core.messages import HumanMessage, SystemMessage
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import time
from io import BytesIO

# --- CONFIGURATION ---
st.set_page_config(
    page_title="Dognosis Financial Report",
    page_icon="🐶",
    layout="wide"
)

# Load API Key from secrets
try:
    API_KEY = st.secrets["GOOGLE_API_KEY"]
except FileNotFoundError:
    st.error("Secrets file not found. Please ensure .streamlit/secrets.toml exists.")
    st.stop()

# --- STYLES ---
st.markdown("""
    <style>
    .stButton>button {
        width: 100%;
        background-color: #1F4E78;
        color: white;
    }
    .metric-card {
        background-color: #f0f2f6;
        padding: 20px;
        border-radius: 10px;
        border-left: 5px solid #1F4E78;
    }
    </style>
""", unsafe_allow_html=True)

# --- HELPER FUNCTIONS ---

def get_llm():
    return ChatGoogleGenerativeAI(
        model="gemini-2.5-flash",
        google_api_key=API_KEY,
        temperature=0.1,
        max_retries=2,
        request_timeout=60
    )

def process_files(kodo_file, trans_file, progress_bar, status_text):
    transactions = []
    errors = []
    
    # 1. LOAD KODO PAY
    try:
        status_text.text("Loading Kodo Pay data...")
        kodo_df = pd.read_csv(kodo_file)
        kodo_debit = kodo_df[kodo_df['Dr/Cr'] == 'Dr'].copy()
        
        for _, row in kodo_debit.iterrows():
            narration = str(row.get('Narration on Kodo Pay', ''))
            comments = str(row.get('Maker Comments', ''))
            
            # FILTER: Skip LQ Prepaid
            if 'lq prepaid' in narration.lower() or 'lq prepaid' in comments.lower():
                continue
                
            transactions.append({
                'source': 'Kodo-Pay',
                'date': str(row.get('Date (IST)', '')),
                'amount': float(row.get('Txn Amount (INR)', 0)),
                'category': str(row.get('Category', 'Uncategorized')),
                'narration': narration,
                'comments': comments,
                'maker_name': str(row.get('Maker Name', '')),
                'description': f"{narration} | {row.get('Category', '')} | {comments} | {row.get('Maker Name', '')}"
            })
    except Exception as e:
        errors.append(f"Kodo Pay Error: {str(e)}")

    # 2. LOAD TRANSACTIONS
    try:
        status_text.text("Loading Transactions data...")
        trans_df = pd.read_csv(trans_file)
        # Filter out FUNDING/CREDIT
        if 'Txn Category' in trans_df.columns:
            trans_debit = trans_df[~trans_df['Txn Category'].isin(['FUNDING', 'CARD_CREDIT'])].copy()
        else:
            trans_debit = trans_df.copy()
            
        for _, row in trans_debit.iterrows():
            merchant = str(row.get('Merchant/Narration', ''))
            
            # FILTER: Skip LQ Prepaid
            if 'lq prepaid' in merchant.lower():
                continue
                
            transactions.append({
                'source': 'Transactions',
                'date': str(row.get('Txn Date', '')),
                'amount': float(row.get('Txn Amount (Rs.)', 0)),
                'category': str(row.get('Expense Category', 'Uncategorized')),
                'merchant': merchant,
                'narration': merchant,
                'comments': str(row.get('Notes', '')),
                'cardholder': f"{row.get('Cardholder First Name', '')} {row.get('Cardholder Last Name', '')}",
                'description': f"{merchant} | {row.get('Expense Category', '')} | {row.get('Notes', '')} | {row.get('Cardholder First Name', '')}"
            })
    except Exception as e:
        errors.append(f"Transactions Error: {str(e)}")

    if not transactions:
        return None, errors

    # 3. CATEGORIZATION
    categorized = []
    llm_needed = []
    
    status_text.text("Applying Business Rules...")
    
    # Business Rules
    for txn in transactions:
        all_text = f"{txn['category']} {txn['description']} {txn.get('narration','')} {txn.get('comments','')}".lower()
        maker_name = txn.get('maker_name', '').lower()
        
        # Rule 1: Mechanical Hardware
        if 'mechanical hardware' in txn['category'].lower() or 'mechanical hardware' in all_text:
            categorized.append({**txn, 'expense_type': 'CAPEX', 'confidence': 'high', 'reasoning': 'Business Rule: Mechanical Hardware', 'method': 'business-rule', 'original_category': txn['category']})
            continue
            
        # Rule 2: Vishwanatha > 1000
        if ('vishwanatha' in maker_name or 'vishwanath' in maker_name) and txn['category'] == 'Uncategorized' and txn['amount'] > 1000:
            categorized.append({**txn, 'expense_type': 'CAPEX', 'confidence': 'high', 'reasoning': 'Business Rule: Vishwanatha > 1000', 'method': 'business-rule', 'category': 'Capital Investment', 'original_category': txn['category']})
            continue
            
        llm_needed.append(txn)

    # LLM Analysis
    if llm_needed:
        llm = get_llm()
        batch_size = 5
        total_batches = (len(llm_needed) + batch_size - 1) // batch_size
        
        system_prompt = """You are a financial analyst for a HARDWARE COMPANY.
CRITICAL BUSINESS RULES:
1. ALL "Mechanical Hardware" is CAPEX.
2. Vishwanatha + Uncategorized + >1000 -> CAPEX.

Categorize as CAPEX or OPEX. If "Uncategorized", assign a proper category.

CAPEX: One-time investments (Equipment, machinery, electronics, IT, construction)
OPEX: Regular/recurring (Rent, utilities, supplies, food, travel, salaries)

Respond with JSON: {"type": "CAPEX" or "OPEX", "category": "assigned category", "reasoning": "brief explanation"}"""

        for i in range(0, len(llm_needed), batch_size):
            batch = llm_needed[i:i+batch_size]
            current_batch = i // batch_size + 1
            
            status_text.text(f"LLM Analyzing batch {current_batch}/{total_batches}...")
            progress_bar.progress(current_batch / total_batches)
            
            batch_prompt = "Categorize:\n\n"
            for idx, txn in enumerate(batch):
                batch_prompt += f"{idx+1}. ₹{txn['amount']:,.2f} | Cat: {txn['category']} | Desc: {txn['description'][:100]}\n"
            
            try:
                response = llm.invoke([SystemMessage(content=system_prompt), HumanMessage(content=batch_prompt)])
                content = response.content.strip().replace('```json', '').replace('```', '').strip()
                results = json.loads(content)
                
                for txn, result in zip(batch, results):
                    assigned_cat = result.get('category', txn['category'])
                    note = f" (LLM assigned: {assigned_cat})" if txn['category'] == 'Uncategorized' and assigned_cat != 'Uncategorized' else ""
                    
                    categorized.append({
                        **txn,
                        'expense_type': result.get('type', 'OPEX'),
                        'category': assigned_cat,
                        'original_category': txn['category'],
                        'confidence': 'llm',
                        'reasoning': f"LLM: {result.get('reasoning', '')}{note}",
                        'method': 'pure-llm'
                    })
            except Exception as e:
                for txn in batch:
                    categorized.append({**txn, 'expense_type': 'OPEX', 'confidence': 'low', 'reasoning': f'Error: {str(e)}', 'method': 'fallback', 'original_category': txn['category']})
            
            time.sleep(1) # Rate limit safety

    return categorized, errors

def create_excel(categorized_data):
    wb = Workbook()
    
    # Styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(size=16, bold=True, color="1F4E78")
    subtitle_font = Font(size=12, bold=True)
    total_font = Font(bold=True, size=11)
    currency_fmt = '₹#,##0.00'
    percent_fmt = '0.0%'
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    capex = [t for t in categorized_data if t['expense_type'] == 'CAPEX']
    opex = [t for t in categorized_data if t['expense_type'] == 'OPEX']
    total_amt = sum(t['amount'] for t in categorized_data)
    capex_amt = sum(t['amount'] for t in capex)
    opex_amt = sum(t['amount'] for t in opex)
    
    # === SHEET 1: EXECUTIVE SUMMARY ===
    ws = wb.active
    ws.title = "Executive Summary"
    
    ws['A1'] = "DOGNOSIS FINANCIAL SPEND REPORT"
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws['A2'].font = Font(size=10, italic=True)
    
    ws['A4'] = "FINANCIAL OVERVIEW"
    ws['A4'].font = subtitle_font
    
    summary_data = [
        ["Metric", "Value"],
        ["Total Transactions", len(categorized_data)],
        ["Total Amount", total_amt],
        ["CAPEX Total", capex_amt],
        ["OPEX Total", opex_amt],
        ["CAPEX %", capex_amt/total_amt if total_amt else 0],
        ["OPEX %", opex_amt/total_amt if total_amt else 0],
    ]
    
    for r, row in enumerate(summary_data, 5):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.border = border
            if r == 5:
                cell.fill = header_fill
                cell.font = header_font
            if c == 2 and r > 6 and r < 10:
                cell.number_format = currency_fmt
            if c == 2 and r >= 10:
                cell.number_format = percent_fmt
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    
    # === SHEET 2: CAPEX STATEMENT ===
    ws_capex = wb.create_sheet("CAPEX Statement")
    
    ws_capex['A1'] = "CAPITAL EXPENDITURES (CAPEX)"
    ws_capex['A1'].font = title_font
    ws_capex.merge_cells('A1:G1')
    
    ws_capex['A2'] = f"Total CAPEX: ₹{capex_amt:,.2f}"
    ws_capex['A2'].font = subtitle_font
    
    headers = ['Date', 'Source', 'Category', 'Amount', 'Description', 'Method', 'Reasoning']
    for c, h in enumerate(headers, 1):
        cell = ws_capex.cell(4, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Group by category for subtotals
    capex_df = pd.DataFrame(capex)
    current_row = 5
    
    if not capex_df.empty:
        for category in sorted(capex_df['category'].unique()):
            cat_txns = capex_df[capex_df['category'] == category]
            
            for _, txn in cat_txns.iterrows():
                ws_capex.cell(current_row, 1, txn['date']).border = border
                ws_capex.cell(current_row, 2, txn['source']).border = border
                ws_capex.cell(current_row, 3, txn['category']).border = border
                cell_amt = ws_capex.cell(current_row, 4, txn['amount'])
                cell_amt.number_format = currency_fmt
                cell_amt.border = border
                ws_capex.cell(current_row, 5, txn['description'][:80]).border = border
                ws_capex.cell(current_row, 6, txn['method']).border = border
                ws_capex.cell(current_row, 7, txn['reasoning'][:80]).border = border
                current_row += 1
            
            # Subtotal
            subtotal = cat_txns['amount'].sum()
            ws_capex.cell(current_row, 3, f"{category} Subtotal").font = total_font
            cell_sub = ws_capex.cell(current_row, 4, subtotal)
            cell_sub.number_format = currency_fmt
            cell_sub.font = total_font
            cell_sub.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            current_row += 1
        
        # Grand Total
        current_row += 1
        ws_capex.cell(current_row, 3, "GRAND TOTAL").font = Font(bold=True, size=12)
        cell_grand = ws_capex.cell(current_row, 4, capex_amt)
        cell_grand.number_format = currency_fmt
        cell_grand.font = Font(bold=True, size=12)
        cell_grand.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    ws_capex.column_dimensions['A'].width = 12
    ws_capex.column_dimensions['B'].width = 15
    ws_capex.column_dimensions['C'].width = 20
    ws_capex.column_dimensions['D'].width = 15
    ws_capex.column_dimensions['E'].width = 40
    ws_capex.column_dimensions['F'].width = 15
    ws_capex.column_dimensions['G'].width = 40
    
    # === SHEET 3: OPEX STATEMENT ===
    ws_opex = wb.create_sheet("OPEX Statement")
    
    ws_opex['A1'] = "OPERATING EXPENDITURES (OPEX)"
    ws_opex['A1'].font = title_font
    ws_opex.merge_cells('A1:G1')
    
    ws_opex['A2'] = f"Total OPEX: ₹{opex_amt:,.2f}"
    ws_opex['A2'].font = subtitle_font
    
    for c, h in enumerate(headers, 1):
        cell = ws_opex.cell(4, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Group by category for subtotals
    opex_df = pd.DataFrame(opex)
    current_row = 5
    
    if not opex_df.empty:
        for category in sorted(opex_df['category'].unique()):
            cat_txns = opex_df[opex_df['category'] == category]
            
            for _, txn in cat_txns.iterrows():
                ws_opex.cell(current_row, 1, txn['date']).border = border
                ws_opex.cell(current_row, 2, txn['source']).border = border
                ws_opex.cell(current_row, 3, txn['category']).border = border
                cell_amt = ws_opex.cell(current_row, 4, txn['amount'])
                cell_amt.number_format = currency_fmt
                cell_amt.border = border
                ws_opex.cell(current_row, 5, txn['description'][:80]).border = border
                ws_opex.cell(current_row, 6, txn['method']).border = border
                ws_opex.cell(current_row, 7, txn['reasoning'][:80]).border = border
                current_row += 1
            
            # Subtotal
            subtotal = cat_txns['amount'].sum()
            ws_opex.cell(current_row, 3, f"{category} Subtotal").font = total_font
            cell_sub = ws_opex.cell(current_row, 4, subtotal)
            cell_sub.number_format = currency_fmt
            cell_sub.font = total_font
            cell_sub.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            current_row += 1
        
        # Grand Total
        current_row += 1
        ws_opex.cell(current_row, 3, "GRAND TOTAL").font = Font(bold=True, size=12)
        cell_grand = ws_opex.cell(current_row, 4, opex_amt)
        cell_grand.number_format = currency_fmt
        cell_grand.font = Font(bold=True, size=12)
        cell_grand.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    ws_opex.column_dimensions['A'].width = 12
    ws_opex.column_dimensions['B'].width = 15
    ws_opex.column_dimensions['C'].width = 20
    ws_opex.column_dimensions['D'].width = 15
    ws_opex.column_dimensions['E'].width = 40
    ws_opex.column_dimensions['F'].width = 15
    ws_opex.column_dimensions['G'].width = 40
    
    # === SHEET 4: CATEGORY BREAKDOWN ===
    ws_cat = wb.create_sheet("Category Breakdown")
    
    ws_cat['A1'] = "SPEND CATEGORY BREAKDOWN"
    ws_cat['A1'].font = title_font
    ws_cat.merge_cells('A1:D1')
    
    # CAPEX Categories
    ws_cat['A3'] = "CAPEX BY CATEGORY"
    ws_cat['A3'].font = subtitle_font
    
    cat_headers = ['Category', 'Amount', '% of CAPEX', 'Transactions']
    for c, h in enumerate(cat_headers, 1):
        cell = ws_cat.cell(4, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    current_row = 5
    if not capex_df.empty:
        capex_by_cat = capex_df.groupby('category').agg({
            'amount': 'sum',
            'date': 'count'
        }).sort_values('amount', ascending=False)
        
        for category, row in capex_by_cat.iterrows():
            ws_cat.cell(current_row, 1, category).border = border
            cell_amt = ws_cat.cell(current_row, 2, row['amount'])
            cell_amt.number_format = currency_fmt
            cell_amt.border = border
            cell_pct = ws_cat.cell(current_row, 3, row['amount']/capex_amt if capex_amt else 0)
            cell_pct.number_format = percent_fmt
            cell_pct.border = border
            ws_cat.cell(current_row, 4, int(row['date'])).border = border
            current_row += 1
    
    # OPEX Categories
    current_row += 2
    ws_cat[f'A{current_row}'] = "OPEX BY CATEGORY"
    ws_cat[f'A{current_row}'].font = subtitle_font
    current_row += 1
    
    for c, h in enumerate(cat_headers, 1):
        cell = ws_cat.cell(current_row, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    current_row += 1
    if not opex_df.empty:
        opex_by_cat = opex_df.groupby('category').agg({
            'amount': 'sum',
            'date': 'count'
        }).sort_values('amount', ascending=False)
        
        for category, row in opex_by_cat.iterrows():
            ws_cat.cell(current_row, 1, category).border = border
            cell_amt = ws_cat.cell(current_row, 2, row['amount'])
            cell_amt.number_format = currency_fmt
            cell_amt.border = border
            cell_pct = ws_cat.cell(current_row, 3, row['amount']/opex_amt if opex_amt else 0)
            cell_pct.number_format = percent_fmt
            cell_pct.border = border
            ws_cat.cell(current_row, 4, int(row['date'])).border = border
            current_row += 1
    
    ws_cat.column_dimensions['A'].width = 30
    ws_cat.column_dimensions['B'].width = 18
    ws_cat.column_dimensions['C'].width = 15
    ws_cat.column_dimensions['D'].width = 15
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# --- MAIN APP ---

def main():
    # Initialize session state
    if 'categorized_data' not in st.session_state:
        st.session_state.categorized_data = None
    
    # Sidebar
    with st.sidebar:
        st.header("Settings")
        st.info("✅ API Key Configured")
        st.info("✅ Hardware Logic Active")
        st.info("✅ Gemini 2.5 Flash")
        
    # Main Content
    st.title("🐶 Dognosis Financial Spend Report")
    st.markdown("Upload your financial sheets to generate the spend report.")
    
    col1, col2 = st.columns(2)
    with col1:
        kodo_file = st.file_uploader("Kodo Pay Reimbursement (CSV)", type=['csv'])
    with col2:
        trans_file = st.file_uploader("Transactions Sheet (CSV)", type=['csv'])
        
    if st.button("Analyze Expenses", type="primary"):
        if not kodo_file or not trans_file:
            st.error("Please upload both CSV files.")
            return
            
        # Progress UI
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        # Run Analysis
        categorized_data, errors = process_files(kodo_file, trans_file, progress_bar, status_text)
        
        if errors:
            for err in errors:
                st.error(err)
        
        if categorized_data:
            # Store in session state
            st.session_state.categorized_data = categorized_data
            status_text.success("Analysis Complete!")
            progress_bar.progress(100)
    
    # Display results if data exists in session state
    if st.session_state.categorized_data:
        categorized_data = st.session_state.categorized_data
        
        # 1. Summary Metrics
        capex = [t for t in categorized_data if t['expense_type'] == 'CAPEX']
        opex = [t for t in categorized_data if t['expense_type'] == 'OPEX']
        
        capex_total = sum(t['amount'] for t in capex)
        opex_total = sum(t['amount'] for t in opex)
        
        st.divider()
        st.subheader("📊 Financial Summary")
        m1, m2, m3 = st.columns(3)
        with m1: st.metric("Total CAPEX", f"₹{capex_total:,.2f}")
        with m2: st.metric("Total OPEX", f"₹{opex_total:,.2f}")
        with m3: st.metric("Transactions", len(categorized_data))
        
        # 2. Download Button (moved here to persist visualizations below)
        st.divider()
        excel_file = create_excel(categorized_data)
        st.download_button(
            label="📥 Download Dognosis Report (4 Sheets)",
            data=excel_file,
            file_name=f"Dognosis_Spend_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )
        
        # 3. Category Analysis (now persistent after download)
        st.divider()
        st.subheader("📈 Category Breakdown")
        
        c1, c2 = st.columns(2)
        
        with c1:
            st.markdown("### CAPEX by Category")
            if capex:
                capex_df = pd.DataFrame(capex)
                capex_chart = capex_df.groupby('category')['amount'].sum().sort_values(ascending=False)
                st.bar_chart(capex_chart)
                st.dataframe(capex_chart.reset_index().rename(columns={'amount': 'Amount (₹)'}), use_container_width=True)
            else:
                st.info("No CAPEX transactions found.")
        
        with c2:
            st.markdown("### OPEX by Category")
            if opex:
                opex_df = pd.DataFrame(opex)
                opex_chart = opex_df.groupby('category')['amount'].sum().sort_values(ascending=False)
                st.bar_chart(opex_chart)
                st.dataframe(opex_chart.reset_index().rename(columns={'amount': 'Amount (₹)'}), use_container_width=True)

if __name__ == "__main__":
    main()
